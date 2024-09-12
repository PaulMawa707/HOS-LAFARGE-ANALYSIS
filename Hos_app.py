import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import zipfile
import shutil
import time

# Set the app layout to wide mode
st.set_page_config(layout="wide")

# Clean the uploaded folder
@st.cache_data(show_spinner=False)
def clean_up_folder(folder_path):
    if os.path.exists(folder_path):
        for _ in range(3):
            try:
                shutil.rmtree(folder_path)
                break
            except PermissionError:
                time.sleep(1)

@st.cache_data(show_spinner=False)
def process_folder(uploaded_folder):
    folder_path = f"temp_folder_{uploaded_folder.name}"
    os.makedirs(folder_path, exist_ok=True)

    with zipfile.ZipFile(uploaded_folder, "r") as zip_ref:
        zip_ref.extractall(folder_path)

    all_data = []

    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            file_path = os.path.join(folder_path, file_name)
            with pd.ExcelFile(file_path) as xls:
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=7)
                    df = df.iloc[:-5, :]
                    df['driver'] = sheet_name
                    site_info = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=6)
                    site_full_text = site_info.iloc[5, 4]
                    site_name = site_full_text.split('Controltech - ')[-1].strip() if 'Controltech - ' in site_full_text else site_full_text
                    df['site'] = site_name
                    all_data.append(df)

    merged_df = pd.concat(all_data, ignore_index=True)
    df_cleaned = merged_df[merged_df['Start Date & Time'].notna() & (merged_df['Start Date & Time'].str.strip() != '')]
    df_cleaned['Start Date & Time'] = pd.to_datetime(df_cleaned['Start Date & Time'], dayfirst=True, errors='coerce')
    df_cleaned['day'] = df_cleaned['Start Date & Time'].dt.day
    df_cleaned['day_name'] = df_cleaned['Start Date & Time'].dt.strftime('%a-%-d')

    selected_columns = ['Start Date & Time', 'day', 'End Date & Time', 'Type', 'Driving Time in Violation (hh:mm:ss)',
                        'Time In Violation (hh:mm:ss)', 'Violation Limit', 'driver', 'site', 'day_name']
    df_cleaned = df_cleaned[selected_columns]
    df_cleaned['Time in Violation (hrs)'] = (
        pd.to_timedelta(df_cleaned['Time In Violation (hh:mm:ss)']).dt.total_seconds() / 3600
    ).round(1)

    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(dataframe_to_rows(df_cleaned, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    clean_up_folder(folder_path)

    return df_cleaned, output

def create_pivot_table(df):
    pivot_table = df.pivot_table(
        index='driver',
        columns='day',
        values='Time in Violation (hrs)',
        aggfunc='sum',
        fill_value=0,
        margins=True,  # Adds totals for rows and columns
        margins_name='Total'  # Name for the totals row/column
    )
    pivot_table.columns = pivot_table.columns.astype(str)
    try:
        pivot_table = pivot_table.sort_index(
            axis=1,
            key=lambda x: pd.to_datetime(
                x.str.split('-').str[1].astype(str) + '-' + x.str.split('-').str[0], errors='coerce'
            )
        )
    except Exception as e:
        st.error(f"Error while sorting columns: {e}")
    pivot_table = pivot_table.round(1)  # Keep the values to one decimal place
    
    return pivot_table


def style_pivot_table(pivot_table):
    def highlight(val):
        if isinstance(val, (int, float)):
            if val > 0:
                return 'background-color: red'
            elif val == 0:
                return 'background-color: green'
        return ''

    styled_table = pivot_table.style.applymap(highlight).format(precision=1)
    return styled_table

def save_pivot_to_excel(pivot_table):
    wb = Workbook()
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=True, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if isinstance(value, (int, float)):
                if value > 0:
                    cell.fill = red_fill
                elif value == 0:
                    cell.fill = green_fill

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def main():
    st.title("HOS LAFARGE ANALYSIS APP")

    uploaded_folder = st.file_uploader(
        "Upload a zip file containing Excel files", 
        type="zip", 
        key="unique_file_uploader_key"
    )

    if uploaded_folder is not None:
        with st.spinner("Processing files..."):
            df_cleaned, result = process_folder(uploaded_folder)
            st.dataframe(df_cleaned)
            st.download_button(
                label="Download Processed Data",
                data=result,
                file_name="df_cleaned_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            site_filter = st.selectbox("Select Site", options=["All"] + sorted(df_cleaned['site'].unique().tolist()))
            type_filter = st.selectbox("Select Type", options=["All"] + sorted(df_cleaned['Type'].unique().tolist()))

            filtered_df = df_cleaned.copy()
            if site_filter != "All":
                filtered_df = filtered_df[filtered_df['site'] == site_filter]
            if type_filter != "All":
                filtered_df = filtered_df[filtered_df['Type'] == type_filter]

            pivot_table = create_pivot_table(filtered_df)

            st.markdown("### Pivot Table")
            styled_pivot_table = style_pivot_table(pivot_table)
            st.dataframe(styled_pivot_table, use_container_width=True)

            pivot_excel = save_pivot_to_excel(pivot_table)
            st.download_button(
                label="Download Pivot Table",
                data=pivot_excel,
                file_name="pivot_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
